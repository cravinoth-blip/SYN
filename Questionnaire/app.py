from flask import Flask, render_template, request, redirect, url_for, g, send_file
import sqlite3
import json
import os
import io
from datetime import datetime

# ── Configuration ─────────────────────────────────────────────────────────────

BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
DATABASE_URL = os.environ.get('DATABASE_URL', '')
USE_POSTGRES = DATABASE_URL.startswith('postgres')

app = Flask(__name__,
            template_folder=os.path.join(BASE_DIR, 'templates'),
            static_folder=os.path.join(BASE_DIR, 'static'))
app.secret_key = os.environ.get('SECRET_KEY', 'bronchiectasis-unlocked-dev')


# ── Database abstraction (SQLite locally, PostgreSQL on Vercel) ───────────────

class DB:
    """Thin wrapper providing a consistent interface over sqlite3 and psycopg2."""

    def __init__(self, conn, pg=False):
        self._c = conn
        self._pg = pg

    def _s(self, sql):
        return sql.replace('?', '%s') if self._pg else sql

    def query(self, sql, params=()):
        """Execute SELECT; return list of dicts."""
        if self._pg:
            import psycopg2.extras
            cur = self._c.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cur.execute(self._s(sql), params)
            return [dict(r) for r in cur.fetchall()]
        else:
            self._c.row_factory = sqlite3.Row
            return [dict(r) for r in self._c.execute(sql, params).fetchall()]

    def one(self, sql, params=()):
        rows = self.query(sql, params)
        return rows[0] if rows else None

    def execute(self, sql, params=()):
        """Execute write statement (no return value)."""
        if self._pg:
            cur = self._c.cursor()
            cur.execute(self._s(sql), params)
            self._c.commit()
        else:
            self._c.execute(sql, params)
            self._c.commit()

    def insert(self, sql, params=()):
        """Execute INSERT and return the new row's id."""
        if self._pg:
            cur = self._c.cursor()
            cur.execute(self._s(sql) + ' RETURNING id', params)
            self._c.commit()
            return cur.fetchone()[0]
        else:
            cur = self._c.execute(sql, params)
            self._c.commit()
            return cur.lastrowid

    def scalar(self, sql, params=()):
        """Execute a scalar query (e.g. COUNT) and return the single value."""
        if self._pg:
            cur = self._c.cursor()
            cur.execute(self._s(sql), params)
            return cur.fetchone()[0]
        else:
            return self._c.execute(sql, params).fetchone()[0]

    def close(self):
        self._c.close()


def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        if USE_POSTGRES:
            import psycopg2
            db = g._database = DB(psycopg2.connect(DATABASE_URL), pg=True)
        else:
            db = g._database = DB(
                sqlite3.connect(os.path.join(BASE_DIR, 'responses.db')), pg=False
            )
    return db


@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()


_SCHEMA_SQLITE = '''
    CREATE TABLE IF NOT EXISTS survey_responses (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        respondent_name TEXT NOT NULL,
        submitted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        q1_patient_count TEXT,
        q2_comorbidities TEXT DEFAULT '{}', q2_etiologies TEXT DEFAULT '{}',
        q2_keine_nicht_erhoben INTEGER DEFAULT 0,
        q3_high_risk TEXT DEFAULT '[]', q3_sonstiges TEXT,
        q4_exaz_0 TEXT, q4_exaz_1 TEXT, q4_exaz_2 TEXT, q4_exaz_more TEXT,
        q5_diagnosis_by TEXT,
        q6_diagnostic TEXT DEFAULT '[]', q6_other TEXT,
        q7_treatment TEXT DEFAULT '[]', q7_other TEXT,
        q8_priorities TEXT DEFAULT '[]', q8_other TEXT,
        q9_unmet_need TEXT, q9_other TEXT,
        q10_monitoring TEXT DEFAULT '[]', q10_other TEXT,
        q11_visit_frequency TEXT, q12_unreported TEXT,
        q13_mdt TEXT DEFAULT '[]', q13_other TEXT,
        q14_guideline TEXT, q14_comment TEXT
    );
    CREATE TABLE IF NOT EXISTS patient_cases (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        survey_response_id INTEGER REFERENCES survey_responses(id),
        case_number INTEGER,
        gender TEXT, age TEXT,
        symptoms TEXT DEFAULT '[]', symptom_severity TEXT,
        comorbidities TEXT DEFAULT '[]', comorbidities_other TEXT,
        fev1_range TEXT, exacerbations TEXT, hospitalizations TEXT,
        base_therapy TEXT DEFAULT '[]', base_therapy_other TEXT,
        acute_therapy TEXT DEFAULT '[]', acute_therapy_other TEXT,
        additional_therapy TEXT DEFAULT '[]', additional_therapy_other TEXT
    );
'''

_SCHEMA_PG = _SCHEMA_SQLITE.replace(
    'INTEGER PRIMARY KEY AUTOINCREMENT', 'SERIAL PRIMARY KEY'
)


def init_db():
    if USE_POSTGRES:
        import psycopg2
        conn = psycopg2.connect(DATABASE_URL)
        cur = conn.cursor()
        # Execute each statement separately (psycopg2 doesn't support executescript)
        for stmt in _SCHEMA_PG.strip().split(';'):
            stmt = stmt.strip()
            if stmt:
                cur.execute(stmt)
        conn.commit()
        conn.close()
    else:
        conn = sqlite3.connect(os.path.join(BASE_DIR, 'responses.db'))
        conn.executescript(_SCHEMA_SQLITE)
        conn.commit()
        conn.close()


# Initialize DB on module load (required for Vercel cold starts)
with app.app_context():
    try:
        init_db()
    except Exception as _e:
        print(f'DB init warning: {_e}')


# ── Helpers ───────────────────────────────────────────────────────────────────

def get_list(form, prefix, keys):
    return [k for k in keys if form.get(f'{prefix}{k}')]


def _parse_json_fields(row):
    """Parse any JSON string values in a dict row."""
    d = dict(row)
    for k, v in d.items():
        if isinstance(v, str) and v and v[0] in ('[', '{'):
            try:
                d[k] = json.loads(v)
            except Exception:
                pass
    return d


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/survey', methods=['GET', 'POST'])
def survey():
    if request.method == 'POST':
        db = get_db()
        f = request.form

        komorbiditat = {}
        for key in ['COPD', 'Asthma', 'NTM', 'Mukoviszidose']:
            if f.get(f'q2_k_{key}'):
                komorbiditat[key] = f.get(f'q2_k_{key}_pct', '')
        if f.get('q2_k_sonstiges'):
            label = f.get('q2_k_sonstiges_label', '').strip() or 'Sonstiges'
            komorbiditat[f'Sonstiges: {label}'] = f.get('q2_k_sonstiges_pct', '')

        aetiologie = {}
        for key in ['idiopathisch', 'COPD', 'Asthma', 'NTM', 'Mukoviszidose']:
            if f.get(f'q2_a_{key}'):
                aetiologie[key] = f.get(f'q2_a_{key}_pct', '')
        if f.get('q2_a_sonstige'):
            label = f.get('q2_a_sonstige_label', '').strip() or 'Sonstige'
            aetiologie[f'Sonstige: {label}'] = f.get('q2_a_sonstige_pct', '')

        q3_keys  = ['exaz_2plus', 'hospitalisierung', 'chron_infektion', 'niedriger_fev1',
                    'progredienter', 'symptombelastung', 'antibiotika']
        q6_keys  = ['sputumkultur', 'lungenfunktion', 'pulsoximetrie', 'feno', 'blutuntersuchungen']
        q7_keys  = ['atemphysio', 'makrolid', 'mukolytika', 'inh_antibiotika',
                    'bronchodilatatoren', 'ics', 'impfungen', 'rehabilitation']
        q8_keys  = ['clearance', 'infektionen', 'exaz_reduktion', 'lebensqualitaet',
                    'lungenfunktion', 'hospitalisierungen', 'komorbiditaeten', 'antibiotika']
        q10_keys = ['lungenfunktion', 'sputumkulturen', 'hrct', 'symptom_lq',
                    'exaz_frequenz', 'physio_technik', 'adhaerenz', 'nach_exaz']
        q13_keys = ['mdt_meetings', 'ad_hoc', 'elektronische_akte',
                    'versorgungspfad', 'informelle_komm', 'keine_zusammenarbeit']

        survey_id = db.insert('''
            INSERT INTO survey_responses (
                respondent_name, q1_patient_count,
                q2_comorbidities, q2_etiologies, q2_keine_nicht_erhoben,
                q3_high_risk, q3_sonstiges,
                q4_exaz_0, q4_exaz_1, q4_exaz_2, q4_exaz_more,
                q5_diagnosis_by, q6_diagnostic, q6_other,
                q7_treatment, q7_other, q8_priorities, q8_other,
                q9_unmet_need, q9_other, q10_monitoring, q10_other,
                q11_visit_frequency, q12_unreported,
                q13_mdt, q13_other, q14_guideline, q14_comment
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''', (
            f.get('respondent_name', ''),
            f.get('q1', ''),
            json.dumps(komorbiditat, ensure_ascii=False),
            json.dumps(aetiologie, ensure_ascii=False),
            1 if f.get('q2_keine') else 0,
            json.dumps(get_list(f, 'q3_', q3_keys), ensure_ascii=False),
            f.get('q3_sonstiges_text', ''),
            f.get('q4_0', ''), f.get('q4_1', ''), f.get('q4_2', ''), f.get('q4_more', ''),
            f.get('q5', ''),
            json.dumps(get_list(f, 'q6_', q6_keys), ensure_ascii=False), f.get('q6_other', ''),
            json.dumps(get_list(f, 'q7_', q7_keys), ensure_ascii=False), f.get('q7_other', ''),
            json.dumps(get_list(f, 'q8_', q8_keys), ensure_ascii=False), f.get('q8_other', ''),
            f.get('q9', ''), f.get('q9_other', ''),
            json.dumps(get_list(f, 'q10_', q10_keys), ensure_ascii=False), f.get('q10_other', ''),
            f.get('q11', ''),
            f.get('q12', ''),
            json.dumps(get_list(f, 'q13_', q13_keys), ensure_ascii=False), f.get('q13_other', ''),
            f.get('q14', ''), f.get('q14_comment', '')
        ))
        return redirect(url_for('cases', survey_id=survey_id))

    return render_template('survey.html')


@app.route('/cases/<int:survey_id>', methods=['GET', 'POST'])
def cases(survey_id):
    if request.method == 'POST':
        db = get_db()
        f = request.form
        symptom_keys     = ['husten', 'auswurf', 'fieber', 'dyspnoe', 'fatigue', 'thoraxschmerzen', 'haemoptysis']
        comorbidity_keys = ['copd', 'asthma', 'ntm']
        base_keys        = ['rauchstopp', 'schutzimpfungen', 'physio', 'sekretolytisch',
                            'vernebler', 'rehabilitation', 'selbstmanagement']
        acute_keys       = ['langzeit_antibiotika', 'antiinflammatorisch', 'thoraxchirurgie',
                            'niv', 'lungentransplantation']
        add_keys         = ['bronchodilatatoren', 'ics', 'mukoaktiv']

        for cn in range(1, 6):
            p = f'c{cn}_'
            db.execute('''
                INSERT INTO patient_cases (
                    survey_response_id, case_number, gender, age,
                    symptoms, symptom_severity, comorbidities, comorbidities_other, fev1_range,
                    exacerbations, hospitalizations,
                    base_therapy, base_therapy_other,
                    acute_therapy, acute_therapy_other,
                    additional_therapy, additional_therapy_other
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            ''', (
                survey_id, cn,
                f.get(f'{p}gender', ''), f.get(f'{p}age', ''),
                json.dumps(get_list(f, f'{p}s_', symptom_keys), ensure_ascii=False),
                f.get(f'{p}severity', ''),
                json.dumps(get_list(f, f'{p}km_', comorbidity_keys), ensure_ascii=False),
                f.get(f'{p}km_andere', ''),
                f.get(f'{p}fev1', ''),
                f.get(f'{p}exaz', ''),
                f.get(f'{p}hosp', ''),
                json.dumps(get_list(f, f'{p}b_', base_keys), ensure_ascii=False),
                f.get(f'{p}b_andere', ''),
                json.dumps(get_list(f, f'{p}a_', acute_keys), ensure_ascii=False),
                f.get(f'{p}a_andere', ''),
                json.dumps(get_list(f, f'{p}z_', add_keys), ensure_ascii=False),
                f.get(f'{p}z_andere', '')
            ))
        return redirect(url_for('success'))

    return render_template('cases.html', survey_id=survey_id)


@app.route('/success')
def success():
    return render_template('success.html')


@app.route('/admin')
def admin():
    db = get_db()
    responses = db.query(
        'SELECT id, respondent_name, submitted_at FROM survey_responses ORDER BY submitted_at DESC'
    )
    count = db.scalar('SELECT COUNT(*) FROM survey_responses')
    return render_template('admin.html', responses=responses, count=count)


@app.route('/admin/<int:response_id>')
def admin_detail(response_id):
    db = get_db()
    response = db.one('SELECT * FROM survey_responses WHERE id = ?', (response_id,))
    if not response:
        return 'Not found', 404
    cases_rows = db.query(
        'SELECT * FROM patient_cases WHERE survey_response_id = ? ORDER BY case_number',
        (response_id,)
    )
    return render_template('admin_detail.html',
                           response=_parse_json_fields(response),
                           cases=[_parse_json_fields(c) for c in cases_rows])


# ── Label mappings for human-readable export ──────────────────────────────────

Q3_LABELS = {
    'exaz_2plus': '≥ 2 Exazerbationen/Jahr', 'hospitalisierung': 'Hospitalisierung',
    'chron_infektion': 'Chronische Infektion', 'niedriger_fev1': 'Niedriger FEV₁',
    'progredienter': 'Progredienter Verlauf', 'symptombelastung': 'Hohe Symptombelastung',
    'antibiotika': 'Häufiger Antibiotikabedarf',
}
Q6_LABELS  = {
    'sputumkultur': 'Sputumkultur/Mikrobiologie', 'lungenfunktion': 'Lungenfunktion (Spirometrie)',
    'pulsoximetrie': 'Pulsoximetrie/BGA', 'feno': 'FeNO',
    'blutuntersuchungen': 'Blutuntersuchungen/Biomarker',
}
Q7_LABELS  = {
    'atemphysio': 'Atemphysiotherapie', 'makrolid': 'Makrolidtherapie',
    'mukolytika': 'Mukolytika/Sekretolytika', 'inh_antibiotika': 'Inhalative Antibiotika',
    'bronchodilatatoren': 'Bronchodilatatoren (LABA/LAMA)', 'ics': 'ICS',
    'impfungen': 'Impfungen', 'rehabilitation': 'Rehabilitation',
}
Q8_LABELS  = {
    'clearance': 'Atemwegs-Clearance', 'infektionen': 'Kontrolle Infektionen',
    'exaz_reduktion': 'Reduktion Exazerbationen', 'lebensqualitaet': 'Lebensqualität',
    'lungenfunktion': 'Erhalt Lungenfunktion', 'hospitalisierungen': 'Vermeidung Hospitalisierungen',
    'komorbiditaeten': 'Management Komorbiditäten', 'antibiotika': 'Minimierung Antibiotika',
}
Q10_LABELS = {
    'lungenfunktion': 'Lungenfunktion', 'sputumkulturen': 'Sputumkulturen', 'hrct': 'HRCT',
    'symptom_lq': 'Symptom/LQ-Bewertung', 'exaz_frequenz': 'Exazerbationsfrequenz',
    'physio_technik': 'Physiotherapie-Technik', 'adhaerenz': 'Adhärenz',
    'nach_exaz': 'Nach-Exazerbations-Reevaluation',
}
Q13_LABELS = {
    'mdt_meetings': 'MDT-Meetings', 'ad_hoc': 'Ad-hoc-Konsultationen',
    'elektronische_akte': 'Elektronische Akte', 'versorgungspfad': 'Versorgungspfad',
    'informelle_komm': 'Informelle Kommunikation', 'keine_zusammenarbeit': 'Keine Zusammenarbeit',
}
BASE_LABELS = {
    'rauchstopp': 'Rauchstopp', 'schutzimpfungen': 'Schutzimpfungen',
    'physio': 'Physiother. Atemtherapie', 'sekretolytisch': 'Sekretolytische Therapie',
    'vernebler': 'Verneblerschulung', 'rehabilitation': 'Rehabilitation',
    'selbstmanagement': 'Selbstmanagement',
}
ACUTE_LABELS = {
    'langzeit_antibiotika': 'Langzeit-Antibiotika', 'antiinflammatorisch': 'Antiinflammatorisch',
    'thoraxchirurgie': 'Thoraxchirurgie', 'niv': 'NIV', 'lungentransplantation': 'Lungentransplantation',
}
ADD_LABELS = {
    'bronchodilatatoren': 'Bronchodilatatoren (LABA/LAMA)', 'ics': 'ICS',
    'mukoaktiv': 'Mukoaktive Langzeittherapie',
}
SYMPTOM_LABELS = {
    'husten': 'Husten', 'auswurf': 'Auswurf', 'fieber': 'Fieber', 'dyspnoe': 'Dyspnoe',
    'fatigue': 'Fatigue', 'thoraxschmerzen': 'Thoraxschmerzen', 'haemoptysis': 'Hämoptysen',
}


def _label_list(raw, mapping):
    if isinstance(raw, str):
        try:
            raw = json.loads(raw)
        except Exception:
            return raw or ''
    if isinstance(raw, list):
        return ', '.join(mapping.get(k, k) for k in raw)
    return ''


def _label_dict(raw):
    if isinstance(raw, str):
        try:
            raw = json.loads(raw)
        except Exception:
            return raw or ''
    if isinstance(raw, dict):
        return ', '.join(f'{k}: {v}%' for k, v in raw.items() if v)
    return ''


def _load_all_data(db, response_id=None):
    if response_id:
        rows = db.query('SELECT * FROM survey_responses WHERE id = ?', (response_id,))
    else:
        rows = db.query('SELECT * FROM survey_responses ORDER BY submitted_at')

    responses = [_parse_json_fields(r) for r in rows]
    cases_by_id = {}
    for r in responses:
        case_rows = db.query(
            'SELECT * FROM patient_cases WHERE survey_response_id = ? ORDER BY case_number',
            (r['id'],)
        )
        cases_by_id[r['id']] = [_parse_json_fields(c) for c in case_rows]
    return responses, cases_by_id


# ── Excel export ──────────────────────────────────────────────────────────────

def _build_excel(responses, cases_by_id):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    PURPLE = 'FF5C2D91'; LIGHT = 'FFEDE5FA'; WHITE = 'FFFFFFFF'
    thin = Side(style='thin', color='FFD0C2E8')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = 'Umfrageantworten'
    survey_cols = [
        ('ID', 'id'), ('Name', 'respondent_name'), ('Eingereicht', 'submitted_at'),
        ('Q1 – Patientenanzahl', 'q1_patient_count'),
        ('Q2 – Komorbiditäten', None), ('Q2 – Ätiologien', None),
        ('Q2 – Keine/nicht erhoben', 'q2_keine_nicht_erhoben'),
        ('Q3 – Hochrisiko-Faktoren', None), ('Q3 – Sonstiges', 'q3_sonstiges'),
        ('Q4 – 0 Exaz %', 'q4_exaz_0'), ('Q4 – 1 Exaz %', 'q4_exaz_1'),
        ('Q4 – 2 Exaz %', 'q4_exaz_2'), ('Q4 – >2 Exaz %', 'q4_exaz_more'),
        ('Q5 – Diagnose durch', 'q5_diagnosis_by'),
        ('Q6 – Diagnostik', None), ('Q6 – Sonstiges', 'q6_other'),
        ('Q7 – Behandlungsansätze', None), ('Q7 – Sonstiges', 'q7_other'),
        ('Q8 – Prioritäten', None), ('Q8 – Sonstiges', 'q8_other'),
        ('Q9 – Ungedeckter Bedarf', 'q9_unmet_need'), ('Q9 – Sonstiges', 'q9_other'),
        ('Q10 – Monitoring', None), ('Q10 – Sonstiges', 'q10_other'),
        ('Q11 – Besuchsfrequenz', 'q11_visit_frequency'),
        ('Q12 – Ungemeldete Exaz.', 'q12_unreported'),
        ('Q13 – MDT', None), ('Q13 – Sonstiges', 'q13_other'),
        ('Q14 – Leitlinienwahrnehmung', 'q14_guideline'), ('Q14 – Kommentar', 'q14_comment'),
    ]
    hf = Font(bold=True, color=WHITE, size=10)
    hfill = PatternFill('solid', fgColor=PURPLE)

    for ci, (h, _) in enumerate(survey_cols, 1):
        cell = ws1.cell(row=1, column=ci, value=h)
        cell.font = hf; cell.fill = hfill
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        cell.border = border
    ws1.row_dimensions[1].height = 30

    for ri, r in enumerate(responses, 2):
        fill = PatternFill('solid', fgColor=LIGHT if ri % 2 == 0 else WHITE)
        vals = [
            r.get('id'), r.get('respondent_name'), r.get('submitted_at'),
            r.get('q1_patient_count'),
            _label_dict(r.get('q2_comorbidities', {})),
            _label_dict(r.get('q2_etiologies', {})),
            'Ja' if r.get('q2_keine_nicht_erhoben') else 'Nein',
            _label_list(r.get('q3_high_risk', []), Q3_LABELS), r.get('q3_sonstiges', ''),
            r.get('q4_exaz_0', ''), r.get('q4_exaz_1', ''),
            r.get('q4_exaz_2', ''), r.get('q4_exaz_more', ''),
            r.get('q5_diagnosis_by', ''),
            _label_list(r.get('q6_diagnostic', []), Q6_LABELS), r.get('q6_other', ''),
            _label_list(r.get('q7_treatment', []), Q7_LABELS), r.get('q7_other', ''),
            _label_list(r.get('q8_priorities', []), Q8_LABELS), r.get('q8_other', ''),
            r.get('q9_unmet_need', ''), r.get('q9_other', ''),
            _label_list(r.get('q10_monitoring', []), Q10_LABELS), r.get('q10_other', ''),
            r.get('q11_visit_frequency', ''), r.get('q12_unreported', ''),
            _label_list(r.get('q13_mdt', []), Q13_LABELS), r.get('q13_other', ''),
            r.get('q14_guideline', ''), r.get('q14_comment', ''),
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws1.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = border

    for ci in range(1, len(survey_cols) + 1):
        ws1.column_dimensions[get_column_letter(ci)].width = 28
    ws1.column_dimensions['A'].width = 6
    ws1.column_dimensions['B'].width = 30
    ws1.column_dimensions['C'].width = 22
    ws1.freeze_panes = 'A2'

    ws2 = wb.create_sheet('Patientenfälle')
    case_cols = ['Respondent', 'Fall Nr.', 'Geschlecht', 'Alter', 'Symptome', 'Schwere',
                 'Komorbiditäten', 'FEV₁', 'Exazerbationen', 'Hospitalisierungen',
                 'Basistherapie', 'Akuttherapie', 'Zusätzliche Therapien']
    for ci, h in enumerate(case_cols, 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.font = hf; cell.fill = hfill
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        cell.border = border
    ws2.row_dimensions[1].height = 30

    ri = 2
    for r in responses:
        for c in cases_by_id.get(r['id'], []):
            fill = PatternFill('solid', fgColor=LIGHT if ri % 2 == 0 else WHITE)
            cm_ = _label_list(c.get('comorbidities', []), {'copd': 'COPD', 'asthma': 'Asthma', 'ntm': 'NTM'})
            if c.get('comorbidities_other'):
                cm_ = (cm_ + ', ' + c['comorbidities_other']).strip(', ')
            vals = [
                r.get('respondent_name', ''), c.get('case_number', ''),
                c.get('gender', ''), c.get('age', ''),
                _label_list(c.get('symptoms', []), SYMPTOM_LABELS),
                c.get('symptom_severity', ''), cm_, c.get('fev1_range', ''),
                c.get('exacerbations', ''), c.get('hospitalizations', ''),
                _label_list(c.get('base_therapy', []), BASE_LABELS) + (
                    (', ' + c['base_therapy_other']) if c.get('base_therapy_other') else ''),
                _label_list(c.get('acute_therapy', []), ACUTE_LABELS) + (
                    (', ' + c['acute_therapy_other']) if c.get('acute_therapy_other') else ''),
                _label_list(c.get('additional_therapy', []), ADD_LABELS) + (
                    (', ' + c['additional_therapy_other']) if c.get('additional_therapy_other') else ''),
            ]
            for ci, val in enumerate(vals, 1):
                cell = ws2.cell(row=ri, column=ci, value=val)
                cell.fill = fill
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = border
            ri += 1

    for ci in range(1, len(case_cols) + 1):
        ws2.column_dimensions[get_column_letter(ci)].width = 28
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 10
    ws2.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── PDF export ────────────────────────────────────────────────────────────────

def _build_pdf(responses, cases_by_id):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, PageBreak
    )

    PURPLE = colors.HexColor('#5C2D91')
    LIGHT  = colors.HexColor('#EDE5FA')
    MID    = colors.HexColor('#9B6FD0')
    GREY   = colors.HexColor('#F4F0FB')

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    s_title  = ParagraphStyle('title',  fontSize=18, textColor=PURPLE, spaceAfter=4,  fontName='Helvetica-Bold')
    s_h1     = ParagraphStyle('h1',     fontSize=13, textColor=PURPLE, spaceAfter=4,  fontName='Helvetica-Bold')
    s_h2     = ParagraphStyle('h2',     fontSize=10, textColor=MID,    spaceAfter=3,  fontName='Helvetica-Bold')
    s_label  = ParagraphStyle('label',  fontSize=8,  textColor=colors.HexColor('#4A3D6A'), spaceAfter=1, fontName='Helvetica-Bold')
    s_value  = ParagraphStyle('value',  fontSize=9,  textColor=colors.black, spaceAfter=6, fontName='Helvetica')
    s_small  = ParagraphStyle('small',  fontSize=7,  textColor=colors.grey, fontName='Helvetica')
    s_footer = ParagraphStyle('footer', fontSize=7,  textColor=colors.grey, fontName='Helvetica', alignment=1)

    story = [
        Spacer(1, 1*cm),
        Paragraph('Bronchiectasis Unlocked', s_title),
        Paragraph('Auswertung – Umfrage für Teilnehmende des Expertenforums', s_h2),
        Paragraph(f'Exportiert am {datetime.now().strftime("%d.%m.%Y %H:%M")}', s_small),
        Spacer(1, 0.4*cm),
        HRFlowable(width='100%', thickness=2, color=PURPLE),
        Spacer(1, 0.3*cm),
        Paragraph(f'{len(responses)} Einträge gespeichert', s_value),
        PageBreak(),
    ]

    grid_style = TableStyle([
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [GREY, colors.white]),
        ('BOX',      (0, 0), (-1, -1), 0.5, colors.HexColor('#D0C2E8')),
        ('INNERGRID',(0, 0), (-1, -1), 0.3, colors.HexColor('#D0C2E8')),
        ('TOPPADDING',    (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING',   (0, 0), (-1, -1), 6),
    ])

    for r in responses:
        story += [
            Paragraph(f'Teilnehmer:in: {r.get("respondent_name","–")}', s_h1),
            Paragraph(f'Eingereicht: {r.get("submitted_at","–")}', s_small),
            Spacer(1, 0.3*cm),
            HRFlowable(width='100%', thickness=1, color=LIGHT),
            Spacer(1, 0.2*cm),
        ]

        q_data = [
            ['Q1 – Patientenanzahl', r.get('q1_patient_count') or '–',
             'Q5 – Diagnose durch',  r.get('q5_diagnosis_by') or '–'],
            ['Q4 – 0 Exaz %', r.get('q4_exaz_0') or '–',
             'Q4 – 1 Exaz %',  r.get('q4_exaz_1') or '–'],
            ['Q4 – 2 Exaz %', r.get('q4_exaz_2') or '–',
             'Q4 – >2 Exaz %', r.get('q4_exaz_more') or '–'],
            ['Q11 – Besuchsfrequenz', r.get('q11_visit_frequency') or '–',
             'Q12 – Ungemeldete Exaz.', r.get('q12_unreported') or '–'],
        ]
        tbl = Table(
            [[Paragraph(c.upper() if i % 2 == 0 else c, s_label if i % 2 == 0 else s_value)
              for i, c in enumerate(row)] for row in q_data],
            colWidths=[4.5*cm, 4*cm, 4.5*cm, 4*cm],
        )
        tbl.setStyle(grid_style)
        story += [tbl, Spacer(1, 0.25*cm)]

        for label, value in [
            ('Q2 – Komorbiditäten',      _label_dict(r.get('q2_comorbidities', {}))),
            ('Q2 – Ätiologien',          _label_dict(r.get('q2_etiologies', {}))),
            ('Q3 – Hochrisiko-Faktoren', _label_list(r.get('q3_high_risk', []), Q3_LABELS) +
             ((' | ' + r['q3_sonstiges']) if r.get('q3_sonstiges') else '')),
            ('Q6 – Diagnostische Maßnahmen', _label_list(r.get('q6_diagnostic', []), Q6_LABELS) +
             ((' | ' + r['q6_other']) if r.get('q6_other') else '')),
            ('Q7 – Behandlungsansätze',  _label_list(r.get('q7_treatment', []), Q7_LABELS) +
             ((' | ' + r['q7_other']) if r.get('q7_other') else '')),
            ('Q8 – Behandlungsprioritäten', _label_list(r.get('q8_priorities', []), Q8_LABELS) +
             ((' | ' + r['q8_other']) if r.get('q8_other') else '')),
            ('Q9 – Ungedeckter Bedarf',  (r.get('q9_unmet_need') or '') +
             ((' | ' + r['q9_other']) if r.get('q9_other') else '')),
            ('Q10 – Monitoring',         _label_list(r.get('q10_monitoring', []), Q10_LABELS) +
             ((' | ' + r['q10_other']) if r.get('q10_other') else '')),
            ('Q13 – MDT-Zusammenarbeit', _label_list(r.get('q13_mdt', []), Q13_LABELS) +
             ((' | ' + r['q13_other']) if r.get('q13_other') else '')),
            ('Q14 – S2k-Leitlinienwahrnehmung', r.get('q14_guideline') or '–'),
        ]:
            story += [Paragraph(label.upper(), s_label), Paragraph(value or '–', s_value)]
        if r.get('q14_comment'):
            story += [Paragraph('Q14 – KOMMENTAR', s_label),
                      Paragraph(r['q14_comment'], s_value)]

        p_cases = cases_by_id.get(r['id'], [])
        if p_cases:
            story += [Spacer(1, 0.2*cm), Paragraph('PATIENTENFÄLLE', s_h2)]
            for c in p_cases:
                story.append(Paragraph(f'Fall {c.get("case_number","")}', s_h2))
                cm_ = _label_list(c.get('comorbidities', []),
                                  {'copd': 'COPD', 'asthma': 'Asthma', 'ntm': 'NTM'})
                if c.get('comorbidities_other'):
                    cm_ = (cm_ + ', ' + c['comorbidities_other']).strip(', ')
                ctbl = Table(
                    [[Paragraph(cell.upper() if i % 2 == 0 else cell,
                                s_label if i % 2 == 0 else s_value)
                      for i, cell in enumerate(row)] for row in [
                        ['Geschlecht / Alter', f'{c.get("gender","–")} / {c.get("age","–")} Jahre',
                         'Schwere', c.get('symptom_severity') or '–'],
                        ['FEV₁', c.get('fev1_range') or '–',
                         'Exazerbationen', c.get('exacerbations') or '–'],
                        ['Hospitalisierungen', c.get('hospitalizations') or '–',
                         'Komorbiditäten', cm_ or '–'],
                    ]],
                    colWidths=[4.5*cm, 4*cm, 4.5*cm, 4*cm],
                )
                ctbl.setStyle(grid_style)
                story += [ctbl, Spacer(1, 0.15*cm)]

                for lbl, val in [
                    ('Symptome',            _label_list(c.get('symptoms', []), SYMPTOM_LABELS) or '–'),
                    ('Basistherapie',        _label_list(c.get('base_therapy', []), BASE_LABELS) +
                     ((' | ' + c['base_therapy_other']) if c.get('base_therapy_other') else '') or '–'),
                    ('Akuttherapie',         _label_list(c.get('acute_therapy', []), ACUTE_LABELS) +
                     ((' | ' + c['acute_therapy_other']) if c.get('acute_therapy_other') else '') or '–'),
                    ('Zusätzliche Therapien', _label_list(c.get('additional_therapy', []), ADD_LABELS) +
                     ((' | ' + c['additional_therapy_other']) if c.get('additional_therapy_other') else '') or '–'),
                ]:
                    story += [Paragraph(lbl.upper(), s_label), Paragraph(val, s_value)]

        story.append(PageBreak())

    if story and isinstance(story[-1], PageBreak):
        story.pop()
    story += [
        Spacer(1, 1*cm),
        HRFlowable(width='100%', thickness=1, color=colors.HexColor('#D0C2E8')),
        Paragraph('Insmed Germany GmbH · The Squaire 12, Am Flughafen, 60549 Frankfurt am Main', s_footer),
    ]
    doc.build(story)
    buf.seek(0)
    return buf


# ── Export routes ─────────────────────────────────────────────────────────────

@app.route('/admin/export/excel')
def export_excel_all():
    db = get_db()
    responses, cases_by_id = _load_all_data(db)
    buf = _build_excel(responses, cases_by_id)
    ts = datetime.now().strftime('%Y%m%d_%H%M')
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'Bronchiektasen_Umfrage_{ts}.xlsx')


@app.route('/admin/export/pdf')
def export_pdf_all():
    db = get_db()
    responses, cases_by_id = _load_all_data(db)
    buf = _build_pdf(responses, cases_by_id)
    ts = datetime.now().strftime('%Y%m%d_%H%M')
    return send_file(buf, mimetype='application/pdf',
                     as_attachment=True, download_name=f'Bronchiektasen_Umfrage_{ts}.pdf')


@app.route('/admin/<int:response_id>/export/excel')
def export_excel_single(response_id):
    db = get_db()
    responses, cases_by_id = _load_all_data(db, response_id)
    if not responses:
        return 'Not found', 404
    buf = _build_excel(responses, cases_by_id)
    name = responses[0]['respondent_name'].replace(' ', '_')[:40]
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'Bronchiektasen_{name}.xlsx')


@app.route('/admin/<int:response_id>/export/pdf')
def export_pdf_single(response_id):
    db = get_db()
    responses, cases_by_id = _load_all_data(db, response_id)
    if not responses:
        return 'Not found', 404
    buf = _build_pdf(responses, cases_by_id)
    name = responses[0]['respondent_name'].replace(' ', '_')[:40]
    return send_file(buf, mimetype='application/pdf',
                     as_attachment=True, download_name=f'Bronchiektasen_{name}.pdf')


if __name__ == '__main__':
    print('Open: http://localhost:5000')
    print('Admin: http://localhost:5000/admin')
    app.run(debug=True, port=5000, host='0.0.0.0')
