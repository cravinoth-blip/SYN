import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Create a new workbook and add sheets
wb = Workbook()

# Cover sheet
cover_sheet = wb.active
cover_sheet.title = "Cover"
cover_sheet.append(["Endometriosis Patient Journey Tracker"])
cover_sheet.append(["Date:", pd.Timestamp.now().strftime('%Y-%m-%d')])
cover_sheet.append(["Contents:"])
cover_sheet.append(["1. Presentation"])
cover_sheet.append(["2. Diagnosis"])
cover_sheet.append(["3. Treatment"])
cover_sheet.append(["4. Re-Diagnosis"])
cover_sheet.append(["5. Tx Adaptation"])
cover_sheet.append(["6. Living With"])

# Define the phases and their data
phases = {
    "Presentation": {
        "headline": "Initial symptoms often misunderstood or dismissed.",
        "feelings": ["Frustration", "Confusion", "Anxiety"],
        "moment_title": "First Doctor Visit",
        "moment_description": "A patient visits her primary care physician complaining of severe menstrual cramps and pelvic pain. She is told it's normal and prescribed painkillers.",
        "emotional_arc": "falling",
        "mindset": "Why is this pain not taken seriously? Am I overreacting?",
        "pain_points": ["Symptoms are often normalized or dismissed by healthcare providers."],
        "stakeholder": "PCP",
        "severity": "high",
        "unmet_needs": ["Better education for PCPs on recognizing endometriosis symptoms."],
        "confidence": "UNSUPPORTED",
        "gaps": ["Lack of specific data on initial presentation experiences."]
    },
    "Diagnosis": {
        "headline": "Diagnosis is often delayed, leading to prolonged suffering.",
        "feelings": ["Desperation", "Relief", "Validation"],
        "moment_title": "Laparoscopic Confirmation",
        "moment_description": "After years of pain, a laparoscopy confirms endometriosis, providing relief but also frustration over the delay.",
        "emotional_arc": "rising",
        "mindset": "Finally, I have an answer. But why did it take so long?",
        "pain_points": ["Delayed diagnosis due to non-specific symptoms and lack of awareness."],
        "stakeholder": "Specialist",
        "severity": "critical",
        "unmet_needs": ["Improved diagnostic pathways and awareness campaigns."],
        "confidence": "UNSUPPORTED",
        "gaps": ["Specific timelines and diagnostic criteria data."]
    },
    "Treatment": {
        "headline": "Treatment options are varied but often have significant side effects.",
        "feelings": ["Hope", "Frustration", "Concern"],
        "moment_title": "Starting Hormonal Therapy",
        "moment_description": "The patient starts on Elagolix, experiencing some relief but also side effects like mood swings.",
        "emotional_arc": "volatile",
        "mindset": "I hope this works, but I'm worried about the side effects.",
        "pain_points": ["Side effects of hormonal treatments can be severe."],
        "stakeholder": "Pharma",
        "severity": "high",
        "unmet_needs": ["Development of treatments with fewer side effects."],
        "confidence": "UNSUPPORTED",
        "gaps": ["Detailed treatment efficacy and side effect profiles."]
    },
    "Re-Diagnosis": {
        "headline": "Re-evaluation often necessary due to persistent symptoms.",
        "feelings": ["Frustration", "Despair", "Determination"],
        "moment_title": "Symptom Recurrence",
        "moment_description": "Despite treatment, symptoms return, prompting further investigation and possible surgery.",
        "emotional_arc": "falling",
        "mindset": "Why are my symptoms back? What else can be done?",
        "pain_points": ["Persistent symptoms despite treatment."],
        "stakeholder": "Specialist",
        "severity": "high",
        "unmet_needs": ["Better long-term management strategies."],
        "confidence": "UNSUPPORTED",
        "gaps": ["Data on re-diagnosis rates and outcomes."]
    },
    "Tx Adaptation": {
        "headline": "Adapting treatment plans to manage side effects and efficacy.",
        "feelings": ["Adaptation", "Hope", "Caution"],
        "moment_title": "Switching Medications",
        "moment_description": "The patient switches from Elagolix to Dienogest after discussing side effects with her doctor.",
        "emotional_arc": "stable",
        "mindset": "I need to find what works best for me.",
        "pain_points": ["Trial and error in finding the right treatment."],
        "stakeholder": "Patient",
        "severity": "moderate",
        "unmet_needs": ["Personalized treatment plans."],
        "confidence": "UNSUPPORTED",
        "gaps": ["Specific adaptation strategies and patient outcomes."]
    },
    "Living With": {
        "headline": "Managing a chronic condition with lifestyle adjustments.",
        "feelings": ["Acceptance", "Empowerment", "Ongoing Concern"],
        "moment_title": "Daily Management",
        "moment_description": "The patient incorporates dietary changes and regular exercise to manage symptoms.",
        "emotional_arc": "rising",
        "mindset": "I can live with this, but I need to stay vigilant.",
        "pain_points": ["Ongoing management requires lifestyle changes."],
        "stakeholder": "Patient",
        "severity": "moderate",
        "unmet_needs": ["Support systems for lifestyle management."],
        "confidence": "UNSUPPORTED",
        "gaps": ["Long-term management and quality of life data."]
    }
}

# Function to create a sheet for each phase
for phase, data in phases.items():
    sheet = wb.create_sheet(title=phase)
    sheet.append(["Headline", "Feelings", "Key Moment Title", "Key Moment Description", "Emotional Arc", "Mindset", "Pain Points", "Stakeholder", "Severity", "Unmet Needs", "Confidence", "Gaps"])
    sheet.append([
        data["headline"], ", ".join(data["feelings"]), data["moment_title"], data["moment_description"], data["emotional_arc"],
        data["mindset"], ", ".join(data["pain_points"]), data["stakeholder"], data["severity"], ", ".join(data["unmet_needs"]),
        data["confidence"], ", ".join(data["gaps"])
    ])

    # Formatting
    for col in range(1, 13):
        column_letter = get_column_letter(col)
        sheet.column_dimensions[column_letter].width = 20
        sheet[f'{column_letter}1'].font = Font(bold=True)

    # Conditional formatting for confidence
    for row in sheet.iter_rows(min_row=2, max_col=12, max_row=2):
        for cell in row:
            if cell.value == "UNSUPPORTED":
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Save the workbook
file_path = "Endometriosis_Patient_Journey_Tracker.xlsx"
wb.save(file_path)
file_path